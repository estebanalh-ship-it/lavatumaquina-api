from functools import wraps
import os
import json
import io
import math
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify
from sqlalchemy import text
from database import engine
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage 
from flask import send_file

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Por favor, inicia sesión para acceder a esta página.', 'warning')
            return redirect(url_for('admin_bp.login'))
        return f(*args, **kwargs)
    return decorated_function

admin_bp = Blueprint('admin', __name__,
                     template_folder='templates',
                     static_folder='static',
                     static_url_path='/admin/static')

@admin_bp.route('/')
@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()

        EMAIL_ADMIN = os.getenv('EMAIL_ADMIN')
        PASSWORD_ADMIN = os.getenv('PASSWORD_ADMIN')

        if email == EMAIL_ADMIN and password == PASSWORD_ADMIN:
            session['admin_logged_in'] = True
            flash('¡Inicio de sesión exitoso!', 'success')
            return redirect(url_for('admin.control_principal'))

        flash('Credenciales inválidas', 'danger')
        return render_template('login.html', error='Credenciales inválidas')

    return render_template('login.html')

@admin_bp.route('/control_principal')
@login_required
def control_principal():
    with engine.connect() as conn:
        total_agendas = conn.execute(text("SELECT COUNT(*) AS c FROM agendas")).scalar()
        total_clientes = conn.execute(text("SELECT COUNT(*) AS c FROM clientes")).scalar()
        clientes = conn.execute(text("""
            SELECT nombre, telefono, email
            FROM clientes
            ORDER BY id_cliente DESC
        """)).mappings().all()
        agendas = conn.execute(text("""
            SELECT
                c.nombre AS cliente,
                a.fecha_agenda AS fecha,
                s.nombre AS servicio,
                s.duracion_min AS duracion
            FROM agendas AS a
            INNER JOIN clientes AS c ON c.id_cliente = a.id_cliente
            INNER JOIN servicios AS s ON s.id_servicio = a.id_servicio
            ORDER BY a.id_agenda DESC
        """)).mappings().all()

    return render_template(
        'control_principal.html',
        total_agendas=total_agendas,
        total_clientes=total_clientes,
        clientes=clientes,
        agendas=agendas
    )
@admin_bp.route('/agenda_manual', methods=['GET', 'POST'])
@login_required
def agenda_manual():
    if request.method == 'POST':
        id_cliente = request.form['id_cliente']
        id_servicio = request.form['id_servicio']
        fecha_agenda = request.form['fecha_agenda']
        hora_agenda = request.form['hora_agenda']
        try:
            with engine.begin() as conn:
                conn.execute(text("""
                    INSERT INTO agendas (id_cliente, id_servicio, fecha_agenda, estado, fecha_creacion)
                    VALUES (:id_cliente, :id_servicio, :fecha_agenda, 'pendiente', NOW())
                """), {
                    'id_cliente': id_cliente,
                    'id_servicio': id_servicio,
                    'fecha_agenda': f"{fecha_agenda} {hora_agenda}"
                })
            flash('Agenda creada correctamente', 'success')
            return redirect(url_for('admin.agenda_manual'))  # Limpia el formulario
        except Exception as e:
            flash(f'Error al crear agenda: {e}', 'danger')
    with engine.connect() as conn:
        clientes = conn.execute(text("SELECT id_cliente, nombre FROM clientes")).mappings().all()
        servicios = conn.execute(text("SELECT id_servicio, nombre, tipo_servicio FROM servicios")).mappings().all()
    return render_template('agenda_manual.html', clientes=clientes, servicios=servicios)

@admin_bp.route('/horas_ocupadas', methods=['GET'])
@login_required
def horas_ocupadas():
    fecha = request.args.get('fecha')
    if not fecha:
        return jsonify([])
    
    try:
        with engine.connect() as conn:
            agendas = conn.execute(text("""
                SELECT DATE_FORMAT(fecha_agenda, '%H:%i') AS hora
                FROM agendas 
                WHERE DATE(fecha_agenda) = :fecha 
                AND estado != 'cancelada'
            """), {'fecha': fecha}).mappings().all()
            bloqueos = conn.execute(text("""
                SELECT DATE_FORMAT(horario_inicio, '%H:%i') AS hora
                FROM bloqueos 
                WHERE fecha = :fecha
            """), {'fecha': fecha}).mappings().all()
            horas_ocupadas_lista = list(set(
                [row['hora'] for row in agendas] + 
                [row['hora'] for row in bloqueos]
            ))            
        return jsonify(horas_ocupadas_lista)      
    except Exception as e:
        print(f"Error al consultar horas ocupadas en admin: {e}")
        return jsonify([])
        
@admin_bp.route('/nuevo_cliente', methods=['GET', 'POST'])
@login_required
def nuevo_cliente():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        rut = request.form.get('rut')
        email = request.form.get('email')
        telefono = request.form.get('telefono')

        if not nombre or not rut or not email or not telefono:
            flash('Por favor completa los campos obligatorios.', 'danger')
            return redirect(url_for('admin.nuevo_cliente'))

        try:
            with engine.begin() as conn:
                conn.execute(text("""
                    INSERT INTO clientes (rut, nombre, apellido, email, telefono, fecha_registro, activo)
                    VALUES (:rut, :nombre, :apellido, :email, :telefono, NOW(), 1)
                """), {
                    'rut': rut,
                    'nombre': nombre,
                    'apellido': apellido,
                    'email': email,
                    'telefono': telefono
                })

            flash('Cliente agregado correctamente.', 'success')
            return redirect(url_for('admin.nuevo_cliente'))
        except Exception as e:
            flash(f'Error al agregar cliente: {e}', 'danger')
            return redirect(url_for('admin.nuevo_cliente'))

    return render_template('nuevo_cliente.html')

@admin_bp.route('/bloqueo_agenda', methods=['GET', 'POST'])
@login_required
def bloqueo_agenda():
    if request.method == 'POST':
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        bandas = request.form.getlist('bandas')
        motivo = request.form.get('motivo', 'Bloqueo por fuerza mayor')

        if not fecha_inicio or not fecha_fin:
            flash('️ Debes seleccionar fecha inicial y final', 'warning')
            return redirect(url_for('admin.bloqueo_agenda'))

        from datetime import datetime, timedelta
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
        
        if fecha_fin_dt < fecha_inicio_dt:
            flash('⚠️ La fecha final no puede ser anterior a la fecha inicial', 'warning')
            return redirect(url_for('admin.bloqueo_agenda'))

        dias_a_bloquear = []
        fecha_actual = fecha_inicio_dt
        while fecha_actual <= fecha_fin_dt:
            dias_a_bloquear.append(fecha_actual.strftime('%Y-%m-%d'))
            fecha_actual += timedelta(days=1)

        diferencia_dias = (fecha_fin_dt - fecha_inicio_dt).days
        
        if diferencia_dias > 0:
            bandas_a_procesar = ['09:00:00', '11:00:00', '15:00:00', '16:00:00']
        else:
            if not bandas:
                flash('⚠️ Debes seleccionar al menos una banda horaria', 'warning')
                return redirect(url_for('admin.bloqueo_agenda'))
            bandas_a_procesar = bandas

        try:
            with engine.begin() as conn:
                contador_bloqueos = 0

                for fecha in dias_a_bloquear:
                    for hora_inicio in bandas_a_procesar:
                        hora_int = int(hora_inicio.split(':')[0])
                        hora_fin = f"{hora_int + 1:02d}:00:00"
                        
                        conn.execute(text("""
                            INSERT INTO bloqueos (fecha, horario_inicio, horario_fin, motivo)
                            VALUES (:fecha, :hora_inicio, :hora_fin, :motivo)
                        """), {
                            'fecha': fecha,
                            'hora_inicio': hora_inicio,
                            'hora_fin': hora_fin,
                            'motivo': motivo
                        })
                        contador_bloqueos += 1

                if diferencia_dias > 0:
                    flash(f'✅ Se bloquearon {len(dias_a_bloquear)} días completos ({contador_bloqueos} bandas en total). Motivo: {motivo}', 'success')
                else:
                    flash(f'✅ Se bloquearon {len(bandas_a_procesar)} bandas horarias para el {fecha_inicio}. Motivo: {motivo}', 'success')
                return redirect(url_for('admin.bloqueo_agenda'))

        except Exception as e:
            flash(f'❌ Error al crear bloqueos: {e}', 'danger')
    return render_template('bloqueo_agenda.html')

@admin_bp.route('/gestion_precios', methods=['GET', 'POST'])
@login_required
def gestion_precios():
    if request.method == 'POST':
        id_servicio = request.form.get('id_servicio')
        precio_nuevo = request.form.get('precio_nuevo')

        try:
            with engine.begin() as conn:
                # Actualizar el precio del servicio
                conn.execute(text("""
                    UPDATE servicios
                    SET precio = :precio_nuevo
                    WHERE id_servicio = :id_servicio
                """), {
                    'precio_nuevo': precio_nuevo,
                    'id_servicio': id_servicio
                })

                # Obtener nombre del servicio para el mensaje
                servicio = conn.execute(text("""
                    SELECT nombre FROM servicios WHERE id_servicio = :id_servicio
                """), {'id_servicio': id_servicio}).fetchone()

                flash(f'✅ Precio actualizado correctamente para {servicio[0]}', 'success')
                return redirect(url_for('admin.gestion_precios'))

        except Exception as e:
            flash(f'❌ Error al actualizar precio: {e}', 'danger')

    with engine.connect() as conn:
        servicios = conn.execute(text("""
            SELECT id_servicio, nombre, tipo_servicio, precio
            FROM servicios
            ORDER BY tipo_servicio, nombre
        """)).mappings().all()

    return render_template('gestion_precios.html', servicios=servicios)

@admin_bp.route('/cotizaciones')
@login_required
def lista_cotizaciones():
    pagina_actual = request.args.get('pagina', 1, type=int)
    por_pagina = 15
    offset = (pagina_actual - 1) * por_pagina
    with engine.connect() as conn:
        total = conn.execute(text(
            "SELECT COUNT(*) FROM cotizaciones"
        )).scalar()
        cotizaciones = conn.execute(text("""
            SELECT id, fecha, nombre_cliente, rut_cliente, total_final, estado 
            FROM cotizaciones 
            ORDER BY id DESC
            LIMIT :limite OFFSET :offset
        """), {'limite': por_pagina, 'offset': offset}).mappings().all()
    total_paginas = math.ceil(total / por_pagina)
    return render_template('cotizaciones_lista.html',
        cotizaciones=cotizaciones,
        pagina_actual=pagina_actual,
        total_paginas=total_paginas
    )
@admin_bp.route('/nueva_cotizacion', methods=['GET', 'POST'])
@login_required
def nueva_cotizacion():
    if request.method == 'POST':
        rut = request.form.get('rut')
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        telefono = request.form.get('telefono', '').strip()
        
        if telefono and not telefono.startswith('+569'):
            telefono = '+569' + telefono
            
        productos = request.form.getlist('items_producto[]')
        cantidades = request.form.getlist('items_cantidad[]')
        precios = request.form.getlist('items_precio[]')
        
        lista_items = []
        total_neto = 0
        
        for prod, cant, prec in zip(productos, cantidades, precios):
            if prod.strip():
                c = float(cant) if cant else 0
                p = float(prec) if prec else 0
                subtotal = c * p
        
                lista_items.append({
                    "producto": prod,
                    "cantidad": c,
                    "precio_unitario": p,
                    "subtotal": subtotal
                })
                total_neto += subtotal

        iva = total_neto * 0.19
        total_final = total_neto + iva
        
        # Redondeo a la decena más cercana como tenía en su lógica original
        total_neto = round(total_neto / 10) * 10
        iva = round(iva / 10) * 10
        total_final = round(total_final / 10) * 10
        
        items_json = json.dumps(lista_items)
        
        try:
            with engine.begin() as conn:
                # 1. Verificar si el cliente ya existe
                cliente_existente = conn.execute(text(
                    "SELECT id_cliente FROM clientes WHERE rut = :rut"
                ), {'rut': rut}).fetchone()
                
                # 2. Si no existe, lo creamos
                if not cliente_existente:
                    conn.execute(text("""
                        INSERT INTO clientes (rut, nombre, email, telefono, fecha_registro, activo)
                        VALUES (:rut, :nombre, :email, :telefono, NOW(), 1)
                    """), {
                        'rut': rut,
                        'nombre': nombre,
                        'email': email,
                        'telefono': telefono
                    })
                
                # 3. Guardar la cotización
                conn.execute(text("""
                    INSERT INTO cotizaciones 
                    (rut_cliente, nombre_cliente, email_cliente, telefono_cliente, 
                     total_neto, iva, total_final, detalle_items, fecha)
                    VALUES (:rut, :nombre, :email, :tel, :neto, :iva, :final, :items, NOW())
                """), {
                    'rut': rut, 
                    'nombre': nombre, 
                    'email': email, 
                    'tel': telefono,
                    'neto': total_neto, 
                    'iva': iva, 
                    'final': total_final, 
                    'items': items_json
                })
                
            flash('✅ Cotización creada con éxito.', 'success')
            return redirect(url_for('admin.lista_cotizaciones'))
            
        except Exception as e:
            flash(f'❌ Error al guardar cotización: {e}', 'danger')

    with engine.connect() as conn:
        clientes = conn.execute(text("""
            SELECT rut, nombre, email, telefono 
            FROM clientes 
            ORDER BY nombre ASC
        """)).mappings().all()
    
    return render_template('cotizaciones_nueva.html', clientes=clientes)

@admin_bp.route('/descargar_cotizacion/<int:id_cotizacion>')
@login_required
def descargar_cotizacion(id_cotizacion):
    """Genera y descarga el archivo Excel de una cotización específica."""
    
    try:
        # 1. Buscar datos en la BD
        with engine.connect() as conn:
            cot = conn.execute(text("SELECT * FROM cotizaciones WHERE id = :id"), 
                               {'id': id_cotizacion}).mappings().fetchone()
        
        if not cot:
            flash('Cotización no encontrada', 'danger')
            return redirect(url_for('admin.lista_cotizaciones'))

        # 2. Recuperar items del JSON
        items = json.loads(cot['detalle_items'])

        # 3. Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Cotizacion_{cot['id']}"

        # Encabezado Empresa
        ws['A1'] = "COMERCIAL Y SERVICIOS INTEGRALES LTM SPA"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = "RUT: 78.290.357-8"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')

        ws['A3'] = "Tel: +569 36473898"
        ws.merge_cells('A3:E3')
        ws['A3'].alignment = Alignment(horizontal='center')

        ws['A4'] = "Emails: lavatumaquina.rengo@gmail.com | vicentealvarado987@gmail.com"
        ws.merge_cells('A4:E4')
        ws['A4'].alignment = Alignment(horizontal='center')

        ws['A5'] = "Dirección: Elicura #375, Rengo, Sexta Región, Chile."
        ws.merge_cells('A5:E5')
        ws['A5'].alignment = Alignment(horizontal='center')

        # --- COMIENZO COTIZACION ---
        ws['A7'] = "COTIZACIÓN DE SERVICIOS"
        ws['A7'].font = Font(bold=True, size=16)
        ws.merge_cells('A7:E7')
        ws['A7'].alignment = Alignment(horizontal='center')

        # Datos del Cliente (bajamos desde fila 9)
        ws['A9']  = "Cliente:";  ws['B9']  = cot['nombre_cliente'] or ""
        ws['A10'] = "RUT:";      ws['B10'] = cot['rut_cliente'] or ""
        ws['A11'] = "Fecha:";    ws['B11'] = str(cot['fecha'])
        ws['C9']  = "Email:";    ws['D9']  = cot['email_cliente'] or ""
        ws['C10'] = "Teléfono:"; ws['D10'] = cot.get('telefono_cliente', '') or ""

        # Encabezados de Tabla
        headers = ["Descripción / Servicio", "Cantidad", "Precio Neto", "IVA (19%)", "Total"]
        ws.append([])       # Espacio (fila 12)
        ws.append(headers)  # Headers (fila 13)
        
        # Capturar la fila donde quedaron los headers
        header_row = ws.max_row  # <-- ✅ Esto guarda el número de fila donde están los encabezados
        
        # Estilo para cabecera
        for col_num in range(1, 6):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = Font(bold=True)
        for item in items:
            # Asegurar que sean números para evitar error matemático
            cant = float(item.get('cantidad', 0))
            precio = float(item.get('precio_unitario', 0))
            subtotal = float(item.get('subtotal', 0))
            
            iva_linea = subtotal * 0.19
            total_linea = subtotal * 1.19
            
            ws.append([
                item.get('producto', ''), 
                cant, 
                precio,
                iva_linea,
                total_linea
            ])
            current_row = ws.max_row # La fila que acabas de escribir
            # Aplicar formato #,##0 a Precio(C), IVA(D) y Total(E)
            ws.cell(row=current_row, column=3).number_format = '#,##0'
            ws.cell(row=current_row, column=4).number_format = '#,##0'
            ws.cell(row=current_row, column=5).number_format = '#,##0'
        # Totales Finales
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        last_item_row = ws.max_row
        for row in range(header_row, last_item_row + 1):
            for col in range(1, 6):  # Columnas A(1) a E(5)
                ws.cell(row=row, column=col).border = thin_border
          
        ws.append([]) # Espacio
        ws.append(["", "", "Total Neto:", int(float(cot['total_neto'] or 0))])
        ws.append(["", "", "IVA (19%):", int(float(cot['iva'] or 0))])
        ws.append(["", "", "TOTAL FINAL:", int(float(cot['total_final'] or 0))])
        ult_fila = ws.max_row
        for r in range(ult_fila-2, ult_fila+1):
            ws.cell(row=r, column=3).font = Font(bold=True) 
            ws.cell(row=r, column=4).number_format = '#,##0' 
        column_widths = {'A': 28, 'B': 18, 'C': 15, 'D': 15, 'E': 15}

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        try:
            from flask import current_app
            img_path = os.path.join(current_app.root_path, 'static', 'cot.png')

            img = ExcelImage(img_path)
            
            img.anchor = f'A{ws.max_row}'
            
            ws.add_image(img)
        except Exception as e:
            print(f"⚠️ Error insertando imagen: {e}")
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nombre_cliente_safe = str(cot['nombre_cliente']).replace(' ', '_')
        nombre_archivo = f"Cotizacion_{cot['id']}_{nombre_cliente_safe}.xlsx"     
        return send_file(
            buffer, 
            as_attachment=True, 
            download_name=nombre_archivo, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"ERROR EXCEL: {e}") 
        flash(f'Error al generar el Excel: {str(e)}', 'danger')
        return redirect(url_for('admin.lista_cotizaciones'))

@admin_bp.route('/descargar_cotizacion_pdf/<int:id_cotizacion>')
@login_required
def descargar_cotizacion_pdf(id_cotizacion):
    """Genera y descarga el archivo PDF de una cotización específica con encabezado corporativo."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.platypus.flowables import HRFlowable
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from flask import current_app
    import io
    import os 
    try:
        with engine.connect() as conn:
            cot = conn.execute(text("SELECT * FROM cotizaciones WHERE id = :id"), 
                               {'id': id_cotizacion}).mappings().fetchone() 
        if not cot:
            flash('Cotización no encontrada', 'danger')
            return redirect(url_for('admin.lista_cotizaciones'))
        items = json.loads(cot['detalle_items'])
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=54, leftMargin=54,
                                topMargin=54, bottomMargin=36)
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
        styles.add(ParagraphStyle(name='EmpresaNombre', 
                                  fontSize=11, 
                                  textColor=colors.HexColor('#1e3a8a'),
                                  fontName='Helvetica-Bold',
                                  alignment=TA_LEFT,
                                  spaceAfter=4))
        styles.add(ParagraphStyle(name='EmpresaDato', 
                                  fontSize=9, 
                                  textColor=colors.HexColor('#334155'),
                                  alignment=TA_LEFT,
                                  spaceAfter=2))
        styles.add(ParagraphStyle(name='TituloCotizacion', 
                                  fontSize=16, 
                                  textColor=colors.HexColor('#1e3a8a'),
                                  fontName='Helvetica-Bold',
                                  alignment=TA_CENTER,
                                  spaceBefore=6,
                                  spaceAfter=6))
        
        logo_path = os.path.join(current_app.root_path, 'static', 'logocot.webp')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=2.2*inch, height=0.8*inch, kind='proportional')
            celda_logo = [[logo]]
        else:
            celda_logo = [[Paragraph("LOGO", styles['Center'])]]
        
        datos_empresa = [
            [Paragraph("COMERCIAL Y SERVICIOS INTEGRALES LTM SPA", styles['EmpresaNombre'])],
            [Paragraph("RUT: 78.290.357-8", styles['EmpresaDato'])],
            [Paragraph("Tel: +569 36473898", styles['EmpresaDato'])],
            [Paragraph("Email: lavatumaquina.rengo@gmail.com", styles['EmpresaDato'])],
            [Paragraph("Dirección: Elicura #375, Rengo, Sexta Región, Chile", styles['EmpresaDato'])]
        ]
        
        tabla_datos = Table(datos_empresa, colWidths=[3.5*inch])
        tabla_datos.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        celda_datos = [[tabla_datos]]
        encabezado = Table([celda_logo[0] + celda_datos[0]], 
                          colWidths=[2.4*inch, 4.6*inch])
        encabezado.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 0),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        elements.append(encabezado)
        
        # Línea separadora azul
        elements.append(HRFlowable(width="100%", thickness=2, 
                                   color=colors.HexColor('#1e3a8a'), 
                                   spaceAfter=8, spaceBefore=4))
        
        # Título de la cotización
        elements.append(Paragraph("COTIZACIÓN DE SERVICIOS", styles['TituloCotizacion']))
        
        # Línea separadora inferior
        elements.append(HRFlowable(width="100%", thickness=1, 
                                   color=colors.HexColor('#94a3b8'), 
                                   spaceAfter=12))
        cliente_data = [
            ['Cliente:', cot['nombre_cliente']],
            ['RUT:', cot['rut_cliente']],
            ['Fecha:', str(cot['fecha'].strftime('%d/%m/%Y') if cot['fecha'] else '')],
            ['Email:', cot['email_cliente'] or ''],
            ['Teléfono:', cot['telefono_cliente'] or '']
        ]
        
        tabla_cliente = Table(cliente_data, colWidths=[1.5*inch, 5.5*inch])
        tabla_cliente.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e3a8a')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        elements.append(tabla_cliente)
        elements.append(Spacer(1, 0.25*inch))
        
        data = [['Descripción / Servicio', 'Cantidad', 'Precio Neto', 'IVA (19%)', 'Total']]
        
        for item in items:
            cant = float(item.get('cantidad', 0))
            precio = float(item.get('precio_unitario', 0))
            subtotal = float(item.get('subtotal', 0))
            iva_linea = subtotal * 0.19
            total_linea = subtotal * 1.19
            
            data.append([
                item.get('producto', ''),
                cant,
                f"${subtotal:,.0f}".replace(',', '.'),
                f"${iva_linea:,.0f}".replace(',', '.'),
                f"${total_linea:,.0f}".replace(',', '.')
            ])
        
        data.append(['', '', 'Total Neto:', f"${int(cot['total_neto'] or 0):,}".replace(',', '.'), ''])
        data.append(['', '', 'IVA (19%):', f"${int(cot['iva'] or 0):,}".replace(',', '.'), ''])
        data.append(['', '', 'TOTAL FINAL:', '', f"${int(cot['total_final'] or 0):,}".replace(',', '.')])
        
        tabla_items = Table(data, colWidths=[2.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.1*inch])
        tabla_items.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -4), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            # Fila de totales en ámbar
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -3), (-1, -1), 10),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#92400e')),
        ]))
        elements.append(tabla_items)
        elements.append(Spacer(1, 0.4*inch))
        elements.append(HRFlowable(width="100%", thickness=1, 
                                   color=colors.HexColor('#94a3b8'), 
                                   spaceAfter=8))
        elements.append(Paragraph("Gracias por preferirnos", styles['Center']))
        elements.append(Paragraph("Esta cotización tiene una validez de 15 días", styles['Center']))

        # 4. Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        # 5. Enviar archivo
        nombre_cliente_safe = str(cot['nombre_cliente']).replace(' ', '_')
        nombre_archivo = f"Cotizacion_{cot['id']}_{nombre_cliente_safe}.pdf"
        
        return send_file(
            buffer, 
            as_attachment=True, 
            download_name=nombre_archivo, 
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"ERROR PDF: {e}")
        flash(f'Error al generar el PDF: {str(e)}', 'danger')
        return redirect(url_for('admin.lista_cotizaciones'))

@admin_bp.route('/editar_cotizacion/<int:id_cotizacion>', methods=['GET', 'POST'])
@login_required
def editar_cotizacion(id_cotizacion):
    if request.method == 'POST':
        rut     = request.form.get('rut')
        nombre  = request.form.get('nombre')
        email   = request.form.get('email')
        telefono = request.form.get('telefono', '').strip()
        if telefono and not telefono.startswith('+569'):
            telefono = '+569' + telefono
        productos  = request.form.getlist('items_producto[]')
        cantidades = request.form.getlist('items_cantidad[]')
        precios    = request.form.getlist('items_precio[]')
        lista_items = []
        total_neto = 0
        for prod, cant, prec in zip(productos, cantidades, precios):
            if prod.strip():
                c = float(cant) if cant else 0
                p = float(prec) if prec else 0
                subtotal = c * p
                lista_items.append({
                    "producto": prod,
                    "cantidad": c,
                    "precio_unitario": p,
                    "subtotal": subtotal
                })
                total_neto += subtotal
        iva         = total_neto * 0.19
        total_final = total_neto + iva
        total_neto = round(total_neto /10) * 10
        iva        = round(iva /10) *10
        total_final = round(total_final /10) *10
        items_json  = json.dumps(lista_items)
        try:
            with engine.begin() as conn:
                conn.execute(text("""
                    UPDATE cotizaciones
                    SET rut_cliente      = :rut,
                        nombre_cliente   = :nombre,
                        email_cliente    = :email,
                        telefono_cliente = :tel,
                        total_neto       = :neto,
                        iva              = :iva,
                        total_final      = :final,
                        detalle_items    = :items
                    WHERE id = :id
                """), {
                    'rut': rut, 'nombre': nombre, 'email': email, 'tel': telefono,
                    'neto': total_neto, 'iva': iva, 'final': total_final,
                    'items': items_json, 'id': id_cotizacion
                })
            flash('✅ Cotización actualizada con éxito.', 'success')
            return redirect(url_for('admin.lista_cotizaciones'))
        except Exception as e:
            flash(f'❌ Error al actualizar: {e}', 'danger')
    with engine.connect() as conn:
        cot = conn.execute(text(
            "SELECT * FROM cotizaciones WHERE id = :id"
        ), {'id': id_cotizacion}).mappings().fetchone()

    if not cot:
        flash('Cotización no encontrada.', 'danger')
        return redirect(url_for('admin.lista_cotizaciones'))
    items = json.loads(cot['detalle_items'])
    return render_template('cotizaciones_editar.html', cot=cot, items=items)
    
@admin_bp.route('/stock_productos')
@login_required
def stock_productos():
    return render_template('stock_productos.html')

@admin_bp.route('/buscar_cliente_admin')
@login_required
def buscar_cliente_admin():
    rut = request.args.get('rut')
    try:
        with engine.connect() as conn:
            cliente = conn.execute(text(
                "SELECT * FROM clientes WHERE rut = :rut"
            ), {'rut': rut}).mappings().fetchone()

        if cliente:
            telefono_bd = str(cliente['telefono']) if cliente['telefono'] else ""
            telefono_limpio = telefono_bd.replace('+569', '')
            return jsonify({
                'existe': True,
                'nombre': cliente['nombre'],
                'email': cliente['email'],
                'telefono': telefono_limpio
            })
        else:
            return jsonify({'existe': False})
    except Exception as e:
        return jsonify({'error': str(e)})

@admin_bp.route('/logout')
@login_required
def logout():
    session.clear()
    flash('Sesión cerrada correctamente', 'success')
    return redirect(url_for('admin_bp.login'))
